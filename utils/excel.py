from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import io


def create_excel_with_styled_header(headers, data_rows):
    wb = Workbook()
    ws = wb.active

    # Style for bold and border
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    vertical_border = Border(left=Side(style="thin"), right=Side(style="thin"))
    bottom_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header with style
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.border = border_style
        cell.alignment = center_alignment
    # Write data rows
    for row_num, row_data in enumerate(data_rows, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = normal_font
            # Apply bottom border to last row
            if row_num == len(data_rows) + 1:
                cell.border = bottom_border
            else:
                cell.border = vertical_border
            cell.alignment = center_alignment

    # Auto-adjust column widths and row heights
    for col_num, header in enumerate(headers, start=1):
        max_length = len(str(header))
        max_lines = 1
        for row_data in data_rows:
            try:
                cell_value = str(row_data[col_num - 1])
                cell_length = len(cell_value)
                cell_lines = cell_value.count("\n") + 1
                if cell_length > max_length:
                    max_length = cell_length
                if cell_lines > max_lines:
                    max_lines = cell_lines
            except IndexError:
                pass
        # Add a little extra space
        adjusted_width = max_length + 6
        ws.column_dimensions[chr(64 + col_num)].width = adjusted_width

        # Adjust row heights for wrapped text
        for row_num in range(1, len(data_rows) + 2):  # +2 for header row
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value and "\n" in str(cell.value):
                ws.row_dimensions[row_num].height = 20 * (
                    str(cell.value).count("\n") + 1
                )

    # Save to in-memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
